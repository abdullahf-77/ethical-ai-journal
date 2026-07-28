"""
Phase 0 — consolidate the 15-domain unified taxonomy into 7 top-level domains.

Reads (READ-ONLY, never modified):
    C:\\Users\\afs12\\Desktop\\taxonomy_completed_unified.xlsx
        sheet "Unified Taxonomy"        -> 15 domains / subdomains / scope / provenance summary
        sheet "Unified Source Mapping"  -> row-level provenance back to the 12 original frameworks

Writes (new files only, inside this project directory):
    taxonomy/taxonomy_v2_7domains.xlsx
    taxonomy/taxonomy_v2_7domains.json
    taxonomy/taxonomy_v2_grouping_rationale.md

Re-runnable: always fully regenerates its three outputs from the source workbook.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SOURCE_XLSX = Path(r"C:\Users\afs12\Desktop\taxonomy_completed_unified.xlsx")
PROJECT_DIR = Path(__file__).resolve().parent
OUT_XLSX = PROJECT_DIR / "taxonomy_v2_7domains.xlsx"
OUT_JSON = PROJECT_DIR / "taxonomy_v2_7domains.json"
OUT_RATIONALE_MD = PROJECT_DIR / "taxonomy_v2_grouping_rationale.md"

TAXONOMY_VERSION = "v2"

# ---------------------------------------------------------------------------
# The 15 -> 7 grouping. This is the one place the mapping is declared; every
# output (xlsx, json, markdown) is derived from this single source of truth.
# ---------------------------------------------------------------------------

TOP_DOMAINS = [
    {
        "id": "TD1",
        "name": "Human Rights, Fairness & Dignity",
        "member_domains": [
            "Human Rights, Dignity & Democratic Values",
            "Fairness, Diversity & Inclusion",
        ],
        "definition": (
            "Whether an AI system respects the fundamental rights, dignity, and equal "
            "treatment of the people and groups it affects. Covers both the foundational "
            "rights/democratic-values register (non-discrimination, dignity, rule of law "
            "as legal and political principles) and its operational counterpart in bias "
            "mitigation, accessibility, and inclusive treatment of protected groups."
        ),
        "rationale": (
            "Both source domains ask the same root question — does this system treat "
            "people justly and respect who they are — from two angles: 'Human Rights, "
            "Dignity & Democratic Values' frames it as a rights/legal obligation, while "
            "'Fairness, Diversity & Inclusion' frames it as a measurable bias/equity "
            "problem (e.g. bias prevention, equitable access, protected-attribute "
            "inference). They are the normative and operational halves of one concern."
        ),
    },
    {
        "id": "TD2",
        "name": "Privacy, Transparency & Explainability",
        "member_domains": [
            "Privacy & Data Protection",
            "Transparency, Explainability & Traceability",
        ],
        "definition": (
            "An individual's or the public's ability to know about, understand, and "
            "control how an AI system collects and uses their information and reaches "
            "its decisions. Spans data protection/confidentiality (what the system knows "
            "and how it's obtained) and system legibility (how the system's behaviour and "
            "outputs can be explained, traced, and disclosed as AI-generated)."
        ),
        "rationale": (
            "Privacy and transparency are both about controlling information flow around "
            "the system, just in opposite directions: privacy restricts what information "
            "flows *into* the system and out to others; transparency/explainability "
            "governs what information flows *out* about how the system works. Many "
            "original subdomains blur the two already (e.g. 'awareness of AI "
            "interaction', 'confidential data in prompt'), so treating them as one "
            "informational-control cluster is a genuine, not arbitrary, merge."
        ),
    },
    {
        "id": "TD3",
        "name": "Security, Resilience & Technical Integrity",
        "member_domains": [
            "Security, Resilience & Technical Abuse",
            "Data, Model & Supply-Chain Integrity",
        ],
        "definition": (
            "The technical robustness of AI systems and their supply chains against "
            "attack, tampering, and failure. Covers adversarial/security threats (attacks, "
            "vulnerabilities, resilience under abuse) together with the integrity of the "
            "data, models, and components feeding the system (poisoning, contamination, "
            "provenance of training data and model artifacts)."
        ),
        "rationale": (
            "Both are 'defend the technical pipeline from compromise' concerns — one from "
            "an external attacker's perspective (security/resilience), the other from a "
            "corrupted-input/corrupted-component perspective (data/model/supply-chain "
            "integrity). Data and model poisoning is itself a named attack vector in the "
            "security literature, so the two domains already overlap mechanically, not "
            "just thematically."
        ),
    },
    {
        "id": "TD4",
        "name": "Safety, Alignment & Human Oversight",
        "member_domains": [
            "Safety, Reliability & Value Alignment",
            "Human Agency, Autonomy & Oversight",
        ],
        "definition": (
            "Whether an AI system behaves reliably, stays aligned with the values and "
            "intentions of its operators and users, and remains under meaningful human "
            "understanding and control. Covers system-level reliability/alignment "
            "(accuracy, dangerous capabilities, value alignment) together with the "
            "human-agency question of who retains the ability to intervene, consent, and "
            "override."
        ),
        "rationale": (
            "A system cannot be judged 'safe' independent of whether a human can still "
            "understand, interrupt, and correct it — the source frameworks pair these "
            "ideas constantly (e.g. 'excessive agency' and 'excessive autonomy' as risks "
            "sit right alongside reliability and value-alignment risks in the same "
            "frameworks). Reliability is the engineering half; human oversight is the "
            "control-authority half of the same failure mode: an unsafe, unaligned, or "
            "unsupervised system."
        ),
    },
    {
        "id": "TD5",
        "name": "Accountability, Governance & Legal Rights",
        "member_domains": [
            "Accountability, Governance & Risk Management",
            "Intellectual Property & Legal Rights",
        ],
        "definition": (
            "The organizational, regulatory, and legal structures that assign "
            "responsibility for an AI system's behaviour and outcomes. Covers "
            "governance/risk-management process (audits, accountability structures, actor "
            "responsibility, risk management) together with the legal-rights regime that "
            "governs what may lawfully be done with data, models, and outputs (copyright, "
            "data-acquisition and transfer restrictions)."
        ),
        "rationale": (
            "Both are 'who is answerable, and under what rulebook' concerns rather than "
            "questions about the system's technical behaviour. Governance defines "
            "internal accountability structures; IP/legal rights define the external "
            "legal constraints those structures must operate inside — the two together "
            "form the full accountability envelope (internal responsibility + external "
            "law) around an AI system."
        ),
    },
    {
        "id": "TD6",
        "name": "Societal, Economic & Environmental Impact",
        "member_domains": [
            "Beneficence, Well-being & Social Good",
            "Socioeconomic & Societal Impact",
            "Sustainability & Environmental Impact",
        ],
        "definition": (
            "AI's broad, often diffuse downstream effects on people, economies, and the "
            "planet, beyond the behaviour of any single system. Covers beneficence and "
            "well-being (does AI actively promote human flourishing and social good), "
            "socioeconomic effects (labour markets, access to essential services, "
            "democratic processes), and environmental impact (energy and resource use, "
            "environmental harm)."
        ),
        "rationale": (
            "These three source domains are the only ones in the taxonomy that are not "
            "about a system's internal behaviour at all — they are about externalities AI "
            "produces on society, the economy, and the environment at large. They scale "
            "at the level of populations and ecosystems rather than individual users, "
            "which is what separates this cluster from, e.g., the individual-rights "
            "cluster in TD1."
        ),
    },
    {
        "id": "TD7",
        "name": "Information & Content Integrity",
        "member_domains": [
            "Information Integrity & Misinformation",
            "Misuse, Manipulation & Harmful Content",
        ],
        "definition": (
            "The integrity of information and content that AI systems produce or "
            "propagate. Covers unintentional harms to the information ecosystem "
            "(disinformation, factual inaccuracy) together with intentional misuse — "
            "manipulation, deepfakes, and generation of harmful or dangerous content."
        ),
        "rationale": (
            "Both domains are about AI corrupting the information ecosystem; the only "
            "difference is intent — misinformation is the unintentional/systemic failure "
            "mode, misuse/manipulation is the deliberate one. Distinguishing them further "
            "would require judging intent, which the source frameworks themselves don't "
            "cleanly separate (e.g. deepfakes appear as both a disinformation and a "
            "misuse concern across frameworks), so keeping them as one top domain avoids "
            "an artificial split."
        ),
    },
]

# sanity: every one of the original 15 must appear in exactly one group
_ALL_ORIGINAL_15 = [
    "Human Rights, Dignity & Democratic Values",
    "Fairness, Diversity & Inclusion",
    "Privacy & Data Protection",
    "Transparency, Explainability & Traceability",
    "Security, Resilience & Technical Abuse",
    "Data, Model & Supply-Chain Integrity",
    "Safety, Reliability & Value Alignment",
    "Human Agency, Autonomy & Oversight",
    "Accountability, Governance & Risk Management",
    "Beneficence, Well-being & Social Good",
    "Socioeconomic & Societal Impact",
    "Sustainability & Environmental Impact",
    "Information Integrity & Misinformation",
    "Misuse, Manipulation & Harmful Content",
    "Intellectual Property & Legal Rights",
]


def domain_to_top_domain_map() -> dict[str, dict]:
    mapping = {}
    for td in TOP_DOMAINS:
        for d in td["member_domains"]:
            mapping[d] = td
    return mapping


def load_source_sheets():
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Read-only source not found: {SOURCE_XLSX}")
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)

    tax_ws = wb["Unified Taxonomy"]
    tax_rows = list(tax_ws.iter_rows(values_only=True))
    tax_header, tax_data = tax_rows[0], tax_rows[1:]

    map_ws = wb["Unified Source Mapping"]
    map_rows = list(map_ws.iter_rows(values_only=True))
    map_header, map_data = map_rows[0], map_rows[1:]

    wb.close()
    return tax_header, tax_data, map_header, map_data


def build_structures():
    tax_header, tax_data, map_header, map_data = load_source_sheets()
    dmap = domain_to_top_domain_map()

    seen_domains = set()
    for row in tax_data:
        seen_domains.add(row[0])
    missing = seen_domains - set(_ALL_ORIGINAL_15)
    extra = set(_ALL_ORIGINAL_15) - seen_domains
    if missing:
        raise ValueError(f"Source workbook has domains not covered by the 7-group mapping: {missing}")
    if extra:
        raise ValueError(f"Mapping references domains not found in the source workbook: {extra}")

    # subdomain rows, annotated with their assigned top domain
    subdomain_rows = []
    for row in tax_data:
        domain = row[0]
        td = dmap[domain]
        subdomain_rows.append(
            {
                "top_domain_id": td["id"],
                "top_domain_name": td["name"],
                "domain": domain,
                "subdomain": row[1],
                "scope": row[2],
                "source_count": row[3],
                "sources": row[4],
                "original_domains": row[5],
                "original_terms": row[6],
            }
        )

    # provenance rows, annotated with their assigned top domain
    provenance_rows = []
    for row in map_data:
        domain = row[0]
        td = dmap[domain]
        provenance_rows.append(
            {
                "top_domain_id": td["id"],
                "top_domain_name": td["name"],
                "domain": domain,
                "subdomain": row[1],
                "source_framework": row[2],
                "original_domain": row[3],
                "original_subdomain": row[4],
                "source_row": row[5],
                "consolidation_action": row[6],
            }
        )

    # mid-tier: the 15 domains with subdomain counts, nested under their top domain
    domain_summary = {}
    for r in subdomain_rows:
        key = r["domain"]
        d = domain_summary.setdefault(
            key,
            {"domain": key, "top_domain_id": r["top_domain_id"], "top_domain_name": r["top_domain_name"], "subdomain_count": 0},
        )
        d["subdomain_count"] += 1

    return subdomain_rows, provenance_rows, domain_summary


def write_json(subdomain_rows, provenance_rows, domain_summary):
    top_domains_out = []
    for td in TOP_DOMAINS:
        domains_out = []
        for dname in td["member_domains"]:
            subs = [
                {
                    "name": r["subdomain"],
                    "scope": r["scope"],
                    "source_count": r["source_count"],
                    "sources": r["sources"],
                    "original_domains": r["original_domains"],
                    "original_terms": r["original_terms"],
                }
                for r in subdomain_rows
                if r["domain"] == dname
            ]
            domains_out.append(
                {
                    "name": dname,
                    "subdomain_count": len(subs),
                    "subdomains": subs,
                }
            )
        top_domains_out.append(
            {
                "id": td["id"],
                "name": td["name"],
                "definition": td["definition"],
                "grouping_rationale": td["rationale"],
                "domain_count": len(domains_out),
                "subdomain_count": sum(d["subdomain_count"] for d in domains_out),
                "domains": domains_out,
            }
        )

    out = {
        "taxonomy_version": TAXONOMY_VERSION,
        "generated_from": str(SOURCE_XLSX),
        "structure": "7 top_domains -> 15 domains -> subdomains",
        "top_domains": top_domains_out,
        "provenance": provenance_rows,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT_JSON}")


def _write_sheet(wb, title, header, rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append(row)
    for c in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 32
    ws.freeze_panes = "A2"
    return ws


def write_xlsx(subdomain_rows, provenance_rows, domain_summary):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_sheet(
        wb,
        "Top Domains",
        ["ID", "Name", "Definition", "Grouping Rationale", "# Domains", "# Subdomains", "Member Domains (15-tier)"],
        [
            (
                td["id"],
                td["name"],
                td["definition"],
                td["rationale"],
                len(td["member_domains"]),
                sum(1 for r in subdomain_rows if r["top_domain_id"] == td["id"]),
                "\n".join(td["member_domains"]),
            )
            for td in TOP_DOMAINS
        ],
    )

    _write_sheet(
        wb,
        "15 Domains (mid-tier)",
        ["Domain", "Top Domain ID", "Top Domain Name", "Subdomain Count"],
        [
            (d["domain"], d["top_domain_id"], d["top_domain_name"], d["subdomain_count"])
            for d in domain_summary.values()
        ],
    )

    _write_sheet(
        wb,
        "Subdomains",
        [
            "Top Domain ID",
            "Top Domain Name",
            "Domain",
            "Subdomain",
            "Domain Scope (Synthesized)",
            "Source Count",
            "Sources",
            "Original Domain(s)",
            "Original Term(s)",
        ],
        [
            (
                r["top_domain_id"],
                r["top_domain_name"],
                r["domain"],
                r["subdomain"],
                r["scope"],
                r["source_count"],
                r["sources"],
                r["original_domains"],
                r["original_terms"],
            )
            for r in subdomain_rows
        ],
    )

    _write_sheet(
        wb,
        "Provenance",
        [
            "Top Domain ID",
            "Top Domain Name",
            "Domain",
            "Subdomain",
            "Source Framework",
            "Original Domain",
            "Original Subdomain",
            "Source Row",
            "Consolidation Action",
        ],
        [
            (
                r["top_domain_id"],
                r["top_domain_name"],
                r["domain"],
                r["subdomain"],
                r["source_framework"],
                r["original_domain"],
                r["original_subdomain"],
                r["source_row"],
                r["consolidation_action"],
            )
            for r in provenance_rows
        ],
    )

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")


def write_rationale_md(domain_summary):
    lines = [
        "# Taxonomy v2 — 7-Domain Grouping Rationale",
        "",
        f"Source: `{SOURCE_XLSX.name}` (sheet \"Unified Taxonomy\", 15 domains / 286 subdomain rows).",
        "",
        "This document explains why each of the 15 unified domains was grouped into one of "
        "7 top-level domains. The grouping is conceptual, not arithmetic — each pair/triad "
        "was merged because the two (or three) domains answer the same underlying question "
        "about an AI system from two complementary angles, not because 15 needed to become 7.",
        "",
    ]
    for td in TOP_DOMAINS:
        lines.append(f"## {td['id']} — {td['name']}")
        lines.append("")
        lines.append(f"**Definition:** {td['definition']}")
        lines.append("")
        lines.append(f"**Member domains (15-tier):** {', '.join(td['member_domains'])}")
        lines.append("")
        sub_n = sum(
            d["subdomain_count"] for d in domain_summary.values() if d["top_domain_id"] == td["id"]
        )
        lines.append(f"**Subdomain count:** {sub_n}")
        lines.append("")
        lines.append(f"**Why these belong together:** {td['rationale']}")
        lines.append("")
    OUT_RATIONALE_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT_RATIONALE_MD}")


def main():
    subdomain_rows, provenance_rows, domain_summary = build_structures()
    write_json(subdomain_rows, provenance_rows, domain_summary)
    write_xlsx(subdomain_rows, provenance_rows, domain_summary)
    write_rationale_md(domain_summary)
    print(f"\nDone. {len(TOP_DOMAINS)} top domains, {len(domain_summary)} mid-tier domains, "
          f"{len(subdomain_rows)} subdomains, {len(provenance_rows)} provenance rows.")


if __name__ == "__main__":
    main()
